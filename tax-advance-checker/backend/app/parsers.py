"""File parsers for trial balance (Excel/PDF) and payslips (PDF).

Standard, pragmatic extraction using pdfplumber + openpyxl/pandas. Anything the
parser can't determine is left for the user to fix in the UI (every field is
editable), per the spec's manual-override requirement.
"""
from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation

import pdfplumber
from openpyxl import load_workbook

import app.enums as e

# --- heuristics for category classification (Hebrew keywords) ---
REVENUE_KW = ["הכנסות", "מכירות", "מחזור", "הכנסה", "תקבולים"]
COGS_KW = ["עלות המכר", "עלות מכר", "קניות", "מלאי", "חומרי גלם", "קבלני משנה"]
EXPENSE_KW = [
    "הוצאות", "שכר", "משכורת", "רכב", "כיבוד", "פחת", "אחזקה", "חשמל",
    "טלפון", "שכירות", "ביטוח", "פרסום", "משרדיות", "עמלות", "ריבית",
]

# default deduction percentages by expense keyword (תיאום מס)
DEDUCTION_DEFAULTS = {
    "רכב": Decimal("45"),
    "כיבוד": Decimal("80"),
    "ביגוד": Decimal("80"),
    "טלפון": Decimal("50"),
    "נייד": Decimal("50"),
    "סלולר": Decimal("50"),
    "תרומות": Decimal("0"),
    "אש\"ל": Decimal("0"),
    "קנסות": Decimal("0"),
}


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal(0)
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return Decimal(0)
    s = str(value).strip().replace(",", "").replace("₪", "").replace("(", "-").replace(")", "")
    if not s or s in ("-", "."):
        return Decimal(0)
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(0)


def classify_category(name: str) -> e.LineCategory:
    # שתי קטגוריות בלבד: הכנסות והוצאות. ברירת מחדל — הוצאה.
    n = name or ""
    if any(k in n for k in REVENUE_KW):
        return e.LineCategory.revenue
    return e.LineCategory.expense


def default_deduction_pct(name: str) -> Decimal:
    for kw, pct in DEDUCTION_DEFAULTS.items():
        if kw in (name or ""):
            return pct
    return Decimal("100")


def _build_line(code, name, debit, credit) -> dict:
    debit_d = _to_decimal(debit)
    credit_d = _to_decimal(credit)
    net = debit_d - credit_d
    category = classify_category(name)
    # revenue is naturally a credit balance → present as positive
    if category == e.LineCategory.revenue:
        net = credit_d - debit_d
    return {
        "account_code": str(code).strip() if code else None,
        "account_name": str(name).strip(),
        "debit_amount": debit_d,
        "credit_amount": credit_d,
        "net_amount": net,
        "category": category,
        "deduction_pct": default_deduction_pct(name),
        "is_manually_overridden": False,
    }


def parse_trial_balance_excel(data: bytes) -> list[dict]:
    """Expect columns roughly: code | name | debit | credit (header row autodetected)."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # find header row (contains a Hebrew name-like column)
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        joined = " ".join(str(c) for c in r if c is not None)
        if any(k in joined for k in ["חשבון", "שם", "חובה", "זכות", "קוד"]):
            header_idx = i
            break

    lines = []
    for r in rows[header_idx + 1:]:
        if r is None or all(c is None for c in r):
            continue
        cells = list(r) + [None] * (4 - len(r)) if len(r) < 4 else list(r)
        code, name = cells[0], cells[1]
        debit, credit = cells[2], cells[3]
        if name is None or str(name).strip() == "":
            continue
        # skip rows where name is actually a number
        if isinstance(name, (int, float)) and code is None:
            continue
        lines.append(_build_line(code, name, debit, credit))
    return lines


_NUM = r"-?[\d,]+\.?\d*"
_LINE_RE = re.compile(rf"^(\d{{3,}})?\s*(.+?)\s+({_NUM})\s+({_NUM})\s*$")


def parse_trial_balance_pdf(data: bytes) -> list[dict]:
    lines = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            # prefer table extraction
            for table in page.extract_tables() or []:
                for row in table:
                    if not row or len(row) < 3:
                        continue
                    cleaned = [c for c in row]
                    # try code,name,debit,credit shape
                    if len(cleaned) >= 4:
                        code, name, debit, credit = cleaned[0], cleaned[1], cleaned[-2], cleaned[-1]
                    else:
                        code, name, debit, credit = None, cleaned[0], cleaned[1], cleaned[2]
                    if not name or not str(name).strip():
                        continue
                    lines.append(_build_line(code, name, debit, credit))
            if lines:
                continue
            # fallback: regex over text lines
            text = page.extract_text() or ""
            for ln in text.splitlines():
                mt = _LINE_RE.match(ln.strip())
                if mt:
                    code, name, debit, credit = mt.groups()
                    lines.append(_build_line(code, name, debit, credit))
    return lines


def parse_trial_balance(filename: str, data: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return parse_trial_balance_excel(data)
    if name.endswith(".pdf"):
        return parse_trial_balance_pdf(data)
    raise ValueError("פורמט קובץ לא נתמך — נדרש Excel או PDF")


# --------------------------------------------------------------------------- #
# Payslip — extract cumulative figures                                         #
# --------------------------------------------------------------------------- #

PAYSLIP_PATTERNS = {
    "gross_cumulative": [r"ברוטו\s*מצטבר", r"סה[\"']?כ\s*ברוטו\s*מצטבר", r"שכר\s*ברוטו\s*מצטבר"],
    "income_tax_cumulative": [r"מס\s*הכנסה\s*מצטבר", r"ניכוי\s*מס\s*מצטבר"],
    "national_insurance_cumulative": [r"ביטוח\s*לאומי\s*מצטבר", r"ב[\"']?ל\s*מצטבר"],
    "health_insurance_cumulative": [r"דמי\s*בריאות\s*מצטבר", r"מס\s*בריאות\s*מצטבר", r"בריאות\s*מצטבר"],
}


def parse_payslip_pdf(data: bytes) -> dict:
    """Best-effort extraction of cumulative payslip figures. Returns whatever it
    finds; the UI lets the user complete/correct any field."""
    result = {k: Decimal(0) for k in PAYSLIP_PATTERNS}
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)

    for field, patterns in PAYSLIP_PATTERNS.items():
        for pat in patterns:
            m = re.search(pat + rf"[^\d\-]*({_NUM})", text)
            if m:
                result[field] = _to_decimal(m.group(1))
                break
    return result
