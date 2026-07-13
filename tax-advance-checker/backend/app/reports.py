"""PDF (ReportLab) and Excel (openpyxl) report generation for a calculation."""
from __future__ import annotations

import io
import os

from bidi.algorithm import get_display
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

FONT_CANDIDATES = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "C:/Windows/Fonts/arial.ttf",
]
_FONT_NAME = "Hebrew"
_font_registered = False


def _ensure_font():
    global _font_registered
    if _font_registered:
        return
    for path in FONT_CANDIDATES:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont(_FONT_NAME, path))
            _font_registered = True
            return
    _FONT_NAME_FALLBACK = "Helvetica"  # noqa
    _font_registered = True


def _rtl(text) -> str:
    return get_display(str(text))


def _color_hex(name: str) -> str:
    return {"green": "#16a34a", "yellow": "#ca8a04", "red": "#dc2626"}.get(name, "#000000")


# --------------------------------------------------------------------------- #
# PDF                                                                          #
# --------------------------------------------------------------------------- #

def generate_pdf(case, result: dict) -> bytes:
    _ensure_font()
    font = _FONT_NAME if _font_registered and os.path.exists(FONT_CANDIDATES[0]) else "Helvetica"
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontName=font, alignment=TA_CENTER, fontSize=16)
    sub = ParagraphStyle("sub", parent=styles["Heading2"], fontName=font, alignment=TA_RIGHT, fontSize=12)
    normal = ParagraphStyle("n", parent=styles["Normal"], fontName=font, alignment=TA_RIGHT, fontSize=10)

    elems = [
        Paragraph(_rtl("דוח בדיקת מקדמות מס וביטוח לאומי"), h),
        Spacer(1, 6 * mm),
        Paragraph(_rtl(f"נישום: {case.taxpayer_name}  |  ת.ז: {case.taxpayer_id_number}"), normal),
        Paragraph(_rtl(f"שנת מס: {case.tax_year}  |  חודשים נבדקים: {case.months_count}"), normal),
        Spacer(1, 6 * mm),
    ]

    def person_section(title, p):
        rows = [
            [_rtl("מדד"), _rtl("ערך")],
            [_rtl("מחזור עסקי"), f"{p['business_revenue']:,.0f}"],
            [_rtl("עלות מכר"), f"{p['business_cogs']:,.0f}"],
            [_rtl("הוצאות חשבונאיות"), f"{p['business_expenses_accounting']:,.0f}"],
            [_rtl("תיאום מס"), f"{p['tax_adjustment']:,.0f}"],
            [_rtl("הכנסה חייבת מעסק"), f"{p['business_taxable_income']:,.0f}"],
            [_rtl("הכנסה חייבת משכר"), f"{p['salary_taxable_income']:,.0f}"],
            [_rtl("הכנסה חייבת כוללת"), f"{p['total_taxable_income']:,.0f}"],
            [_rtl("נקודות זיכוי"), f"{p['credit_points_total']:.2f}"],
            [_rtl("מס צפוי (אחרי זיכויים)"), f"{p['income_tax_after_credits']:,.0f}"],
            [_rtl("מס ששולם"), f"{p['income_tax_paid']:,.0f}"],
            [_rtl("פער מס"), f"{p['income_tax_gap']:,.0f}"],
            [_rtl("כיסוי מס %"), f"{p['income_tax_coverage_pct']:.1f}%"],
            [_rtl("ביטוח לאומי צפוי"), f"{p['ni_expected']:,.0f}"],
            [_rtl("דמי בריאות צפוי"), f"{p['ni_health_expected']:,.0f}"],
            [_rtl("ב\"ל ובריאות ששולם"), f"{p['ni_paid']:,.0f}"],
            [_rtl("פער ב\"ל"), f"{p['ni_gap']:,.0f}"],
            [_rtl("מקדמה חודשית מומלצת"), f"{p['ni_monthly_recommended']:,.0f}"],
            [_rtl("פער כולל"), f"{p['total_gap']:,.0f}"],
        ]
        t = Table(rows, colWidths=[90 * mm, 60 * mm], hAlign="RIGHT")
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(_color_hex(p.get("score_color", "")))),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ]))
        return [Paragraph(_rtl(title), sub), Spacer(1, 2 * mm), t, Spacer(1, 6 * mm)]

    elems += person_section("נישום ראשי", result["primary"])
    if result.get("spouse"):
        elems += person_section("בן/בת זוג", result["spouse"])
    if result.get("combined"):
        c = result["combined"]
        rows = [
            [_rtl("תמונת מצב משותפת"), ""],
            [_rtl("מחזור משותף"), f"{c['combined_revenue']:,.0f}"],
            [_rtl("מס צפוי כולל"), f"{c['income_tax_expected']:,.0f}"],
            [_rtl("מס ששולם כולל"), f"{c['income_tax_paid']:,.0f}"],
            [_rtl("ב\"ל צפוי כולל"), f"{c['ni_expected']:,.0f}"],
            [_rtl("ב\"ל ששולם כולל"), f"{c['ni_paid']:,.0f}"],
            [_rtl("פער כולל"), f"{c['total_gap']:,.0f}"],
        ]
        t = Table(rows, colWidths=[90 * mm, 60 * mm], hAlign="RIGHT")
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(_color_hex(c.get("score_color", "")))),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ]))
        elems += [t]

    doc.build(elems)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Excel                                                                        #
# --------------------------------------------------------------------------- #

def generate_excel(case, result: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "סיכום"
    ws.sheet_view.rightToLeft = True

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    white_bold = Font(bold=True, color="FFFFFF")
    right = Alignment(horizontal="right")

    ws["A1"] = "דוח בדיקת מקדמות מס וביטוח לאומי"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"נישום: {case.taxpayer_name}   ת.ז: {case.taxpayer_id_number}"
    ws["A3"] = f"שנת מס: {case.tax_year}   חודשים: {case.months_count}"

    def write_person(ws, start_row, title, p):
        r = start_row
        ws.cell(r, 1, title).font = bold
        r += 1
        for cell in (ws.cell(r, 1, "מדד"), ws.cell(r, 2, "ערך")):
            cell.font = white_bold
            cell.fill = header_fill
        r += 1
        fields = [
            ("מחזור עסקי", p["business_revenue"]),
            ("עלות מכר", p["business_cogs"]),
            ("הוצאות חשבונאיות", p["business_expenses_accounting"]),
            ("הוצאה מותרת", p["business_expenses_allowed"]),
            ("תיאום מס", p["tax_adjustment"]),
            ("הכנסה חייבת מעסק", p["business_taxable_income"]),
            ("הכנסה חייבת משכר", p["salary_taxable_income"]),
            ("הכנסה חייבת כוללת", p["total_taxable_income"]),
            ("נקודות זיכוי", p["credit_points_total"]),
            ("שווי זיכויים", p["credit_points_value_ils"]),
            ("מס צפוי", p["income_tax_after_credits"]),
            ("מס ששולם", p["income_tax_paid"]),
            ("פער מס", p["income_tax_gap"]),
            ("כיסוי מס %", p["income_tax_coverage_pct"]),
            ("ביטוח לאומי צפוי", p["ni_expected"]),
            ("דמי בריאות צפוי", p["ni_health_expected"]),
            ("ב\"ל ובריאות ששולם", p["ni_paid"]),
            ("פער ב\"ל", p["ni_gap"]),
            ("מקדמה חודשית מומלצת", p["ni_monthly_recommended"]),
            ("פער כולל", p["total_gap"]),
            ("ציון", p.get("score_color", "")),
        ]
        for label, value in fields:
            ws.cell(r, 1, label).alignment = right
            ws.cell(r, 2, value)
            r += 1
        return r + 1

    row = 5
    row = write_person(ws, row, "נישום ראשי", result["primary"])
    if result.get("spouse"):
        row = write_person(ws, row, "בן/בת זוג", result["spouse"])
    if result.get("combined"):
        c = result["combined"]
        ws.cell(row, 1, "תמונת מצב משותפת").font = bold
        row += 1
        for label, value in [
            ("מחזור משותף", c["combined_revenue"]),
            ("מס צפוי כולל", c["income_tax_expected"]),
            ("מס ששולם כולל", c["income_tax_paid"]),
            ("ב\"ל צפוי כולל", c["ni_expected"]),
            ("ב\"ל ששולם כולל", c["ni_paid"]),
            ("פער כולל", c["total_gap"]),
        ]:
            ws.cell(row, 1, label).alignment = right
            ws.cell(row, 2, value)
            row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
